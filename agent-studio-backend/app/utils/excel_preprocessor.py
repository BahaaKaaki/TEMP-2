"""
Smart Excel preprocessor for unstructured spreadsheets.

Handles common issues in real-world Excel files:
- Detects actual header rows (skipping titles, sources, legends)
- Filters out non-data sheets (cover pages, navigation, indices)
- Resolves merged cells with forward-filling
- Trims trailing empty columns
- Cleans and normalizes data for structured ingestion

NOTE: Uses openpyxl with read_only=False to access merged cell ranges.
This loads the full workbook into memory, so a file-size guard is applied;
files above MAX_PREPROCESS_BYTES fall back to the naive (read_only) parser
in schema_inference.
"""

import io
import logging
import time
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

MIN_DATA_ROWS = 2
MIN_DATA_COLS = 2
MAX_HEADER_SCAN_ROWS = 30
YEAR_RANGE = (1900, 2100)

# Files above this size skip the smart preprocessor (merged-cell resolution
# requires read_only=False which loads the full workbook into memory).
MAX_PREPROCESS_BYTES = 50 * 1024 * 1024  # 50 MB


def preprocess_excel(file_content: bytes, file_name: str) -> List[dict]:
    """
    Parse an Excel file with smart header detection and sheet filtering.

    Returns a list of dicts (one per viable data sheet) matching the
    schema_inference.parse_excel_data output format:
        {file_name, sheet_name, headers, rows, total_rows, sample_rows}

    Returns an empty list (triggering the naive fallback) if the file
    exceeds MAX_PREPROCESS_BYTES.
    """
    if len(file_content) > MAX_PREPROCESS_BYTES:
        logger.info(
            "File %s is %.1f MB — skipping smart preprocessor (limit %.0f MB)",
            file_name,
            len(file_content) / (1024 * 1024),
            MAX_PREPROCESS_BYTES / (1024 * 1024),
        )
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed; cannot preprocess Excel files")
        return []

    t0 = time.monotonic()
    wb = None
    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        results: List[dict] = []
        for sheet in wb.worksheets:
            try:
                sheet_data = _process_sheet(sheet, file_name)
                if sheet_data:
                    results.append(sheet_data)
            except Exception:
                logger.warning(
                    "Skipping sheet '%s' in %s due to processing error",
                    sheet.title, file_name, exc_info=True,
                )

        elapsed = time.monotonic() - t0
        logger.info(
            "Preprocessed %s: %d/%d sheets viable in %.2fs",
            file_name, len(results), len(wb.worksheets), elapsed,
        )
        return results

    except Exception as e:
        logger.warning("Failed to load Excel file %s: %s", file_name, e)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Sheet processing
# ---------------------------------------------------------------------------

def _process_sheet(sheet, file_name: str) -> Optional[dict]:
    """Process a single sheet: detect headers, clean data, check viability."""
    sheet_name = sheet.title

    merged_map = _build_merged_cell_map(sheet)
    raw_rows = _read_rows_with_merges(sheet, merged_map)

    if not raw_rows or len(raw_rows) < MIN_DATA_ROWS + 1:
        logger.debug("Sheet '%s' skipped: too few rows (%d)", sheet_name, len(raw_rows))
        return None

    header_idx = _find_header_row(raw_rows)
    if header_idx is None:
        logger.debug("Sheet '%s' skipped: no viable header row found", sheet_name)
        return None

    headers_raw = raw_rows[header_idx]
    data_raw = raw_rows[header_idx + 1:]

    col_count = _effective_column_count(headers_raw, data_raw)
    if col_count < MIN_DATA_COLS:
        logger.debug("Sheet '%s' skipped: only %d columns", sheet_name, col_count)
        return None

    headers = _clean_headers(headers_raw[:col_count])

    data_rows: List[List[str]] = []
    for row in data_raw:
        cells = [_cell_to_str(row[i]) if i < len(row) else "" for i in range(col_count)]
        if any(c.strip() for c in cells):
            data_rows.append(cells)

    if len(data_rows) < MIN_DATA_ROWS:
        logger.debug("Sheet '%s' skipped: only %d data rows", sheet_name, len(data_rows))
        return None

    data_rows = _forward_fill_groups(headers, data_rows)
    sample_rows = data_rows[:20]

    return {
        "file_name": file_name,
        "sheet_name": sheet_name,
        "headers": headers,
        "rows": data_rows,
        "total_rows": len(data_rows),
        "sample_rows": sample_rows,
    }


# ---------------------------------------------------------------------------
# Merged cell handling
# ---------------------------------------------------------------------------

def _build_merged_cell_map(sheet) -> Dict[Tuple[int, int], Any]:
    """Map every cell in a merged range to the top-left cell's value."""
    merged_map: Dict[Tuple[int, int], Any] = {}
    for merge_range in sheet.merged_cells.ranges:
        value = sheet.cell(row=merge_range.min_row, column=merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged_map[(row, col)] = value
    return merged_map


def _read_rows_with_merges(sheet, merged_map: Dict) -> List[List]:
    """Read all rows, substituting merged-cell references with their source value."""
    rows: List[List] = []
    for row in sheet.iter_rows(values_only=False):
        cells: List = []
        for cell in row:
            key = (cell.row, cell.column)
            cells.append(merged_map.get(key, cell.value))
        rows.append(cells)
    return rows


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _find_header_row(rows: List[List]) -> Optional[int]:
    """
    Score each candidate row and return the index of the most likely header.

    Scoring weights:
      - fill_ratio   (0.25): headers tend to have most cells populated
      - string_ratio (0.30): header cells are text (or year-like numbers), not data
      - width_score  (0.20): header width should match the table's column span
      - unique_ratio (0.15): header values should be mostly distinct
      - position     (0.10): prefer rows closer to the top
    """
    scan_limit = min(len(rows), MAX_HEADER_SCAN_ROWS)

    # Use the maximum column span across the sheet (rightmost non-empty column)
    # instead of median fill count — data rows may be sparse but still span wide.
    table_span = _max_column_span(rows)

    best_score = -1.0
    best_idx: Optional[int] = None

    for i in range(scan_limit):
        row = rows[i]
        non_empty = [c for c in row if c is not None and str(c).strip()]
        fill_count = len(non_empty)

        if fill_count < MIN_DATA_COLS:
            continue

        total = max(len(row), 1)
        fill_ratio = fill_count / total

        string_count = sum(1 for c in non_empty if _is_header_like(c))
        string_ratio = string_count / max(fill_count, 1)

        if table_span > 0:
            width_score = min(fill_count, table_span) / max(fill_count, table_span)
        else:
            width_score = fill_ratio

        str_vals = [str(c).strip().lower() for c in non_empty]
        unique_ratio = len(set(str_vals)) / max(len(str_vals), 1)

        # There must be real data rows below.  We only require MIN_DATA_COLS
        # filled cells per row (not a fraction of fill_count) because data
        # tables can be very sparse while still having a fully-populated header.
        rows_below = rows[i + 1 : i + 11]
        data_below_count = sum(
            1
            for r in rows_below
            if sum(1 for c in r if c is not None and str(c).strip()) >= MIN_DATA_COLS
        )
        if data_below_count < min(3, len(rows_below)):
            continue

        # Reject ultra-sparse candidates: if row fills less than 20% of the
        # table span, it's probably a title or nav bar, not a real header.
        if table_span > 0 and fill_count / table_span < 0.20:
            continue

        score = (
            fill_ratio * 0.25
            + string_ratio * 0.30
            + width_score * 0.20
            + unique_ratio * 0.15
            + (1.0 - i / max(scan_limit, 1)) * 0.10
        )

        if score > best_score:
            best_score = score
            best_idx = i

    if best_score < 0.25:
        return None

    return best_idx


def _is_header_like(value) -> bool:
    """Return True if the value looks like a column header rather than a data point."""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    try:
        n = int(float(s))
        return YEAR_RANGE[0] <= n <= YEAR_RANGE[1]
    except (ValueError, TypeError):
        pass
    try:
        float(s)
        return False
    except (ValueError, TypeError):
        pass
    return True


# ---------------------------------------------------------------------------
# Column / header helpers
# ---------------------------------------------------------------------------

def _max_column_span(rows: List[List]) -> int:
    """Rightmost non-empty column index across all rows (approximates table width)."""
    span = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i] is not None and str(row[i]).strip():
                span = max(span, i + 1)
                break
    return span


def _effective_column_count(headers: List, data_rows: List[List]) -> int:
    """Rightmost column index (1-based) that has any non-empty value in headers or data."""
    max_col = 0
    for i in range(len(headers) - 1, -1, -1):
        if headers[i] is not None and str(headers[i]).strip():
            max_col = i + 1
            break
    for row in data_rows[:50]:
        for i in range(len(row) - 1, -1, -1):
            if row[i] is not None and str(row[i]).strip():
                max_col = max(max_col, i + 1)
                break
    return max_col


def _clean_headers(raw_headers: List) -> List[str]:
    """Convert raw header values to unique, non-empty strings."""
    headers: List[str] = []
    seen: Dict[str, int] = {}

    for val in raw_headers:
        h = str(val).strip() if val is not None else ""
        if not h:
            h = f"column_{len(headers) + 1}"

        # Integers that look like years — keep as-is
        try:
            n = int(float(h))
            if YEAR_RANGE[0] <= n <= YEAR_RANGE[1]:
                h = str(n)
        except (ValueError, TypeError):
            pass

        base = h
        if h in seen:
            seen[h] += 1
            h = f"{base}_{seen[h]}"
        else:
            seen[h] = 0

        headers.append(h)

    return headers


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# Forward-fill grouped columns
# ---------------------------------------------------------------------------

def _forward_fill_groups(
    headers: List[str], data_rows: List[List[str]]
) -> List[List[str]]:
    """
    Forward-fill the first few text-heavy columns that look like group labels
    (e.g. Region, Category).  Only acts when 30-90 % of cells in a column are
    empty — the signature of vertically-merged group cells in Excel.
    """
    if not data_rows or not headers:
        return data_rows

    max_group_cols = min(3, len(headers))

    for col_idx in range(max_group_cols):
        values = [row[col_idx] if col_idx < len(row) else "" for row in data_rows]
        total = len(values)
        if total == 0:
            continue

        empty_ratio = sum(1 for v in values if not v.strip()) / total

        if not (0.3 <= empty_ratio <= 0.9):
            continue

        non_empty = [v for v in values if v.strip()]
        text_ratio = sum(1 for v in non_empty if not _is_pure_number(v)) / max(len(non_empty), 1)
        if text_ratio > 0.5:
            last = ""
            for row in data_rows:
                if col_idx < len(row):
                    if row[col_idx].strip():
                        last = row[col_idx]
                    else:
                        row[col_idx] = last

    return data_rows


def _is_pure_number(s: str) -> bool:
    try:
        float(s.replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False
