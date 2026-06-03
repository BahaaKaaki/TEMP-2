"""
Schema inference for CSV/Excel structured data in knowledge bases.

Infers column types and suggests descriptions for tables to support
SQL generation and data understanding.
"""

import csv
import io
import json
import logging
import re
from datetime import datetime
from typing import List, Optional

logger = logging.getLogger(__name__)

SCHEMA_INFER_BINDING = "service.schema_infer"

# Date formats to try (in detection priority order)
DATE_FORMATS = [
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%Y/%m/%d",
    "%m-%d-%Y",
    "%d.%m.%Y",
    "%Y.%m.%d",
    "%B %d, %Y",  # January 01, 2025
    "%b %d, %Y",  # Jan 01, 2025
    "%d %B %Y",   # 01 January 2025
]

DATETIME_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%m/%d/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%m/%d/%Y %H:%M",
]

BOOLEAN_TRUE = frozenset({"true", "yes", "1", "y", "on"})
BOOLEAN_FALSE = frozenset({"false", "no", "0", "n", "off"})


def _is_empty(value: str) -> bool:
    """Check if a value is considered empty/null."""
    if value is None:
        return True
    s = str(value).strip().lower()
    return s in ("", "null", "none", "na", "n/a", "-")


def _try_bool(value: str) -> Optional[bool]:
    """Try to parse as boolean. Returns None if not a boolean."""
    s = str(value).strip().lower()
    if s in BOOLEAN_TRUE:
        return True
    if s in BOOLEAN_FALSE:
        return False
    return None


def _try_int(value: str) -> Optional[int]:
    """Try to parse as integer. Returns None if not an integer."""
    s = str(value).strip()
    if not s:
        return None
    try:
        # Reject floats
        if "." in s or "e" in s.lower():
            return None
        return int(s)
    except ValueError:
        return None


def _try_numeric(value: str) -> Optional[float]:
    """Try to parse as numeric (float). Returns None if not numeric."""
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _try_date(value: str) -> Optional[datetime]:
    """Try to parse as date only. Returns None if not a date."""
    s = str(value).strip()
    if not s or len(s) > 50:
        return None
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(s, fmt)
            return dt
        except ValueError:
            continue
    return None


def _try_datetime(value: str) -> Optional[datetime]:
    """Try to parse as datetime. Returns None if not a datetime."""
    s = str(value).strip()
    if not s or len(s) > 80:
        return None
    for fmt in DATETIME_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Fallback to date formats (date implies datetime at midnight)
    return _try_date(s)


def infer_column_types(headers: List[str], sample_rows: List[List[str]]) -> List[dict]:
    """
    Infer data types for each column from sample values.

    Type detection priority: boolean -> integer -> numeric -> date -> datetime -> text.
    A column is nullable if any sample value is empty/null.

    Returns:
        List of dicts: {"column_name": str, "data_type": str, "nullable": bool}
        Data types: "text", "integer", "numeric", "date", "datetime", "boolean"
    """
    if not headers:
        return []

    num_cols = len(headers)
    results: List[dict] = []

    for col_idx, col_name in enumerate(headers):
        values = []
        for row in sample_rows:
            if col_idx < len(row):
                val = row[col_idx]
                values.append(val if val is not None else "")

        nullable = any(_is_empty(v) for v in values)
        non_empty = [str(v).strip() for v in values if not _is_empty(v)]

        if not non_empty:
            results.append({
                "column_name": col_name,
                "data_type": "text",
                "nullable": True,
            })
            continue

        # Type detection with priority: boolean -> integer -> numeric -> date -> datetime -> text
        data_type = "text"

        # Boolean: all non-empty values must parse as bool
        if all(_try_bool(v) is not None for v in non_empty):
            data_type = "boolean"
        # Integer: all non-empty values must parse as int (and not float)
        elif all(_try_int(v) is not None for v in non_empty):
            data_type = "integer"
        # Numeric: all non-empty values must parse as float
        elif all(_try_numeric(v) is not None for v in non_empty):
            data_type = "numeric"
        # Date: all non-empty values must parse as date (no time component)
        elif all(_try_date(v) is not None for v in non_empty):
            data_type = "date"
        # Datetime: all non-empty values must parse as datetime
        elif all(_try_datetime(v) is not None for v in non_empty):
            data_type = "datetime"

        results.append({
            "column_name": col_name,
            "data_type": data_type,
            "nullable": nullable,
        })

    return results


def _parse_json_response(content: str) -> dict:
    """Extract JSON from LLM response, handling markdown fences."""
    match = re.search(r"```(?:json)?\s*\n?(.*?)\n?```", content, re.DOTALL)
    if match:
        content = match.group(1).strip()
    try:
        data = json.loads(content)
        if isinstance(data, dict):
            return data
    except json.JSONDecodeError:
        pass
    brace = content.find("{")
    if brace != -1:
        depth, start = 0, brace
        for i in range(start, len(content)):
            if content[i] == "{":
                depth += 1
            elif content[i] == "}":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(content[start : i + 1])
                    except json.JSONDecodeError:
                        break
    return {}


def _get_schema_llm():
    """Get a LangChain chat model for schema description inference."""
    from app.config.llm_config import LLMClientManager
    return LLMClientManager.get_client_for_binding(
        SCHEMA_INFER_BINDING,
        temperature=0.0,
        max_tokens=1024,
    )


SCHEMA_DESCRIBE_SYSTEM = """\
You are a data analyst. You will receive column names and a sample of rows \
from a structured file (CSV or Excel sheet).

Your task:
1. Write a **table_description** (1-2 sentences): what this dataset represents, \
its business domain, and what kind of questions it could answer.
2. For **every** column, write a short description (1 sentence max) that \
explains the semantic meaning of that column — not just its data type. \
An LLM will use these descriptions to decide which columns to query.

Rules:
- Use the EXACT column names as JSON keys (case-sensitive).
- Be specific: "Total revenue from the sale in USD" is better than "revenue value".
- If a column is an ID or code, say what entity it identifies.
- Respond with ONLY a JSON object (no markdown fences, no extra text):
{"table_description": "...", "column_descriptions": {"col_name": "...", ...}}"""


async def suggest_descriptions(
    headers: List[str],
    sample_rows: List[List[str]],
    file_name: str,
) -> dict:
    """
    Use an LLM to suggest a table description and per-column descriptions.

    Returns:
        {"table_description": str, "column_descriptions": {"col_name": "description", ...}}
    """
    from langchain_core.messages import HumanMessage, SystemMessage

    sample = sample_rows[:10]
    header_line = " | ".join(headers)
    rows_text = "\n".join(
        " | ".join(str(cell) for cell in row) for row in sample
    )

    user_prompt = (
        f"File: {file_name}\n\n"
        f"{header_line}\n"
        f"{rows_text}\n\n"
        f"Columns ({len(headers)}): {', '.join(headers)}"
    )

    try:
        llm = _get_schema_llm()
        response = await llm.ainvoke([
            SystemMessage(content=SCHEMA_DESCRIBE_SYSTEM),
            HumanMessage(content=user_prompt),
        ])
        content = response.content
        if not content:
            return {"table_description": "", "column_descriptions": {}}

        raw = _parse_json_response(content.strip())
        table_desc = raw.get("table_description", "") or ""
        col_descs = raw.get("column_descriptions", {})
        if not isinstance(col_descs, dict):
            col_descs = {}

        return {
            "table_description": table_desc,
            "column_descriptions": col_descs,
        }
    except Exception as exc:
        logger.warning("Schema description LLM call failed: %s", exc)
        return {"table_description": "", "column_descriptions": {}}


def parse_csv_data(file_content: bytes, file_name: str) -> dict:
    """
    Parse CSV bytes into headers and rows.

    Uses csv.Sniffer for delimiter detection. Returns first 20 rows as sample_rows.
    """
    try:
        text = file_content.decode("utf-8", errors="replace")
    except Exception as e:
        logger.warning("Failed to decode CSV as UTF-8: %s", e)
        text = file_content.decode("latin-1", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)

    if not rows:
        return {
            "file_name": file_name,
            "headers": [],
            "rows": [],
            "total_rows": 0,
            "sample_rows": [],
        }

    headers = rows[0]
    data_rows = [r for r in rows[1:] if any(cell.strip() for cell in r)]
    sample_rows = data_rows[:20]

    return {
        "file_name": file_name,
        "headers": headers,
        "rows": data_rows,
        "total_rows": len(data_rows),
        "sample_rows": sample_rows,
    }


def parse_excel_data(file_content: bytes, file_name: str) -> List[dict]:
    """
    Parse Excel bytes into per-sheet data.

    Uses a smart preprocessor that automatically:
    - Detects the real header row (skipping titles, sources, legends)
    - Filters out non-data sheets (cover pages, navigation, indices)
    - Resolves merged cells and forward-fills group columns
    - Trims trailing empty columns

    Falls back to naive row-0-as-header parsing if the preprocessor fails.

    Returns a list of dicts (one per sheet), each with:
        file_name, sheet_name, headers, rows, total_rows, sample_rows
    """
    from app.utils.excel_preprocessor import preprocess_excel

    try:
        result = preprocess_excel(file_content, file_name)
        if result:
            return result
        logger.info(
            "Smart preprocessor returned no sheets for %s; falling back to naive parse",
            file_name,
        )
    except Exception as e:
        logger.warning(
            "Smart preprocessor failed for %s (%s); falling back to naive parse",
            file_name, e,
        )

    return _parse_excel_naive(file_content, file_name)


def _parse_excel_naive(file_content: bytes, file_name: str) -> List[dict]:
    """Original naive parser: row 0 = headers, rest = data. Used as fallback."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed; cannot parse Excel files")
        return []

    try:
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to load Excel file %s: %s", file_name, e)
        return []

    result: List[dict] = []
    for sheet in wb.worksheets:
        sheet_name = sheet.title
        raw_rows = list(sheet.iter_rows(values_only=True))

        if not raw_rows:
            continue

        headers = ["" if v is None else str(v) for v in raw_rows[0]]

        data_rows = []
        for row in raw_rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            data_rows.append(["" if v is None else str(v) for v in row])

        if not data_rows:
            continue

        sample_rows = data_rows[:20]

        result.append({
            "file_name": file_name,
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": data_rows,
            "total_rows": len(data_rows),
            "sample_rows": sample_rows,
        })

    try:
        wb.close()
    except Exception:
        pass

    return result
